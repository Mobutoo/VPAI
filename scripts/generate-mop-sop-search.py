#!/usr/bin/env python3
"""
Génère docs/mop-sop-search.xlsx — Moteur de recherche MOP/SOP NOC
Sources : mop-route.json, mop-incidents-alignment.yml, mop-get.json
Plan   : .planning/notes/plan-mop-search-engine.md
Date   : 2026-04-14
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================
# DATA
# ============================================================

# Mapping anciens IDs SOP → nouveaux IDs (fusion 13→9)
MAPPING_OLD_NEW = {
    'SOP-01': 'SOP-01',
    'SOP-02': 'SOP-02',
    'SOP-03': 'SOP-03',
    'SOP-04': 'SOP-04',
    'SOP-05': 'SOP-04',   # fusionné dans SOP-04
    'SOP-06': 'SOP-06',
    'SOP-07': 'SOP-06',   # fusionné dans SOP-06
    'SOP-08': 'SOP-01',   # OOB Cogent → dans SOP-01
    'SOP-09': 'SOP-09',
    'SOP-10': 'SOP-07',   # OLP → nouveau SOP-07
    'SOP-11': 'SOP-05',   # PM counters → nouveau SOP-05
    'SOP-12': 'SOP-05',   # fusionné dans SOP-05
    'SOP-13': 'SOP-08',   # Viavi → nouveau SOP-08
}

# 9 SOPs cibles
SOP_NEW = {
    'SOP-01': ('S0.10', 'Vérification OOB & accès management'),
    'SOP-02': ('S0.11', 'Reset GCC / OSPF Ribbon'),
    'SOP-03': ('S0.12', 'Escalade Ribbon TAC'),
    'SOP-04': ('S0.13', 'Diagnostic signal OTN (OSNR + FEC/BER)'),
    'SOP-05': ('S0.14', 'PM Counters OTN'),
    'SOP-06': ('S0.15', 'Remplacement matériel (carte + SFP)'),
    'SOP-07': ('S0.16', 'Bascule OLP'),
    'SOP-08': ('S0.17', 'Validation Viavi post-restauration'),
    'SOP-09': ('S0.18', 'Diagnostic firewall Lugos'),
}

# Famille MOP par domaine
FAMILLE_MOP = {
    'SUPERVISION':      ('M0.20', 'OPSSD-03.03.M0.20-P-V1', 'MOP Supervision OTN/OOB'),
    'PHOTONIQUE':       ('M0.21', 'OPSSD-03.03.M0.21-P-V1', 'MOP Photonique OTN'),
    'CARTE SERVICE':    ('M0.22', 'OPSSD-03.03.M0.22-P-V1', 'MOP Cartes OTN'),
    'CARTE MANAGEMENT': ('M0.22', 'OPSSD-03.03.M0.22-P-V1', 'MOP Cartes OTN'),
    'CARTE INFRA':      ('M0.22', 'OPSSD-03.03.M0.22-P-V1', 'MOP Cartes OTN'),
    'CHASSIS':          ('M0.23', 'OPSSD-03.03.M0.23-P-V1', 'MOP Châssis OTN'),
    'PORT NNI':         ('M0.24', 'OPSSD-03.03.M0.24-P-V1', 'MOP Ports NNI/UNI OTN'),
    'PORT UNI':         ('M0.24', 'OPSSD-03.03.M0.24-P-V1', 'MOP Ports NNI/UNI OTN'),
}

# 40 incidents (id, symptome, domaine, criticite, sops_anciens)
INCIDENTS = [
    # SUPERVISION (9)
    ('S1',  'NE et PHO passent bleu/gris sur Lightsoft',                         'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-03 → SOP-02'),
    ('S2',  "Pas d'accès OOB depuis l'extérieur (perte lien L3 Cogent)",         'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-02 → SOP-08 → SOP-11'),
    ('S3',  'Coupure backbone Cogent (lien L2)',                                  'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-02 → SOP-08 → SOP-11'),
    ('S4',  'Cogent prend le relai — perte lien L2 EuNetwork',                   'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-02 → SOP-08 → SOP-11'),
    ('S5',  'Plus aucun lien supervision externe (double OOB simultané down)',    'SUPERVISION',      'CRITIQUE', 'SOP-01 → SOP-02 → SOP-08 → SOP-09 → SOP-11'),
    ('S6',  'Impossible de se connecter à Lightsoft',                             'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-03 → SOP-11'),
    ('S7',  'Points gris et bleu Lightsoft — chassis non joignables',             'SUPERVISION',      'MAJEUR',   'SOP-01 → SOP-03 → SOP-06 → SOP-10 → SOP-11'),
    ('S8',  'Perte supervision — certificat SSL expiré',                          'SUPERVISION',      'MINEUR',   'SOP-01 → SOP-03 → SOP-10'),
    ('S9',  "Alarme MAJOR 'No OSPF hello packets received' — auto-cleared <30s", 'SUPERVISION',      'MINEUR',   'SOP-01 → SOP-04 → SOP-03'),
    # PHOTONIQUE (5)
    ('P1',  'Alarme LOS sur OSC + ampli (perte lien opérateur franche)',          'PHOTONIQUE',       'CRITIQUE', 'SOP-01 → SOP-04 → SOP-08 → SOP-11'),
    ('P2',  'Alarme SPAN LOSS sur PHO — valeur dB dégradée',                     'PHOTONIQUE',       'MAJEUR',   'SOP-01 → SOP-04 → SOP-07 → SOP-11'),
    ('P3',  'Alarme OLP HS — plus de bascule possible',                           'PHOTONIQUE',       'CRITIQUE', 'SOP-01 → SOP-04 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('P4',  'Vizee indique dégradation signal opérateur (faux positif SFP OSC)',  'PHOTONIQUE',       'MAJEUR',   'SOP-01 → SOP-04 → SOP-05 → SOP-10 → SOP-11'),
    ('P5',  'LOS/LOF multi-alarmes + absence totale signal (coupure fibre)',      'PHOTONIQUE',       'CRITIQUE', 'SOP-01 → SOP-04 → SOP-07 → SOP-08 → SOP-11 → SOP-13'),
    # CARTE SERVICE (3)
    ('CS1', 'Alarmes multiples — perte tous ports LB (carte LB défectueuse)',     'CARTE SERVICE',    'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CS2', 'Alarmes multiples — perte cascade LB (carte HB défectueuse)',        'CARTE SERVICE',    'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CS3', 'Alarmes multiples — cascade HB + LB (carte VHB défectueuse)',        'CARTE SERVICE',    'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    # CARTE MANAGEMENT (3)
    ('CM1', 'Chassis gris/bleu sur Lightsoft — perte RCP chassis 9608',           'CARTE MANAGEMENT', 'MAJEUR',   'SOP-01 → SOP-03 → SOP-06 → SOP-10 → SOP-11'),
    ('CM2', 'Chassis 9603 gris/bleu — lien opérateur potentiellement perdu',      'CARTE MANAGEMENT', 'MAJEUR',   'SOP-01 → SOP-03 → SOP-06 → SOP-10 → SOP-11'),
    ('CM3', 'Chassis vert et joignable — alarme fichier système RCP 9608',        'CARTE MANAGEMENT', 'MAJEUR',   'SOP-01 → SOP-03 → SOP-06 → SOP-10'),
    # CARTE INFRA (5)
    ('CI1', 'Alarme température chassis (perte carte FAN 9608)',                  'CARTE INFRA',      'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CI2', 'Alarme FAN + RCP inaccessible (switch intégré FAN 9608)',            'CARTE INFRA',      'MAJEUR',   'SOP-01 → SOP-05 → SOP-06 → SOP-10 → SOP-11'),
    ('CI3', 'Alarme alimentation chassis (perte carte Power)',                    'CARTE INFRA',      'MAJEUR',   'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CI4', 'Alarme rectifier — 1 voie restante (chassis 9608)',                  'CARTE INFRA',      'MAJEUR',   'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CI5', 'Extinction totale du rack — silence alarmes (perte secteur)',        'CARTE INFRA',      'CRITIQUE', 'SOP-01 → SOP-05 → SOP-09 → SOP-11'),
    # CHASSIS (5)
    ('CH1', 'Alarme générale + perte tous ports LB (chassis LB)',                 'CHASSIS',          'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CH2', 'Alarme générale + cascade LB (chassis HB)',                          'CHASSIS',          'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CH3', 'Alarme générale + cascade HB + LB (chassis VHB)',                    'CHASSIS',          'CRITIQUE', 'SOP-01 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CH4', 'Alarmes LOS OSC + ampli — chassis 9603 sans OLP',                   'CHASSIS',          'CRITIQUE', 'SOP-01 → SOP-04 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    ('CH5', 'Alarmes LOS trails — tous clients impactés (chassis 9603 avec OLP)', 'CHASSIS',          'CRITIQUE', 'SOP-01 → SOP-04 → SOP-05 → SOP-10 → SOP-11 → SOP-13'),
    # PORT NNI (9)
    ('N1',  'Alarmes multiples port NNI 16 LB',                                  'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N2',  'Alarmes multiples port NNI 17 LB',                                  'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N3',  'Alarmes NNI 16 HB + cascade LB',                                    'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N4',  'Alarmes NNI 20 HB + cascade LB',                                    'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N5',  'Alarmes NNI 0 HB — tous ports HB impactés',                          'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N6',  'Alarmes NNI 1 HB — tous ports HB impactés',                          'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N7',  'Alarmes NNI 12 VHB + cascade HB + LB',                              'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N8',  'Alarmes NNI 13 VHB + cascade HB + LB',                              'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    ('N9',  'Alarmes NNI 0/1 VHB + cascade NNI 12/13',                           'PORT NNI',         'CRITIQUE', 'SOP-01 → SOP-07 → SOP-10 → SOP-11 → SOP-13'),
    # PORT UNI (1)
    ('U1',  'Alarme sur port UNI',                                                'PORT UNI',         'MAJEUR',   'SOP-01 → SOP-07 → SOP-10'),
]

# ============================================================
# HELPERS
# ============================================================

def convert_sops_anciens(old_chain: str) -> list:
    """Convertit chaîne SOPs anciens → liste nouveaux IDs dédupliqués (ordre préservé)."""
    parts = [p.strip() for p in old_chain.split('→')]
    seen = set()
    result = []
    for p in parts:
        if p in MAPPING_OLD_NEW:
            new = MAPPING_OLD_NEW[p]
            if new not in seen:
                result.append(new)
                seen.add(new)
    return result


def build_db_rows():
    """Construit les lignes de la feuille DB."""
    rows = []
    for inc_id, symptome, domaine, criticite, sops_anciens in INCIDENTS:
        famille_code, ref_mop, titre_mop = FAMILLE_MOP[domaine]
        new_sops = convert_sops_anciens(sops_anciens)
        sops_nouveaux_str = ' → '.join(new_sops)
        sops_refs = ' → '.join(SOP_NEW[s][0] for s in new_sops)
        sops_titres = ' | '.join(SOP_NEW[s][1] for s in new_sops)
        nb_sops = len(new_sops)
        rows.append([
            inc_id,          # A: ID
            symptome,        # B: Alarme_Lightsoft
            domaine,         # C: Domaine
            criticite,       # D: Criticite
            famille_code,    # E: Famille_MOP (M0.20…)
            ref_mop,         # F: Ref_MOP
            titre_mop,       # G: Titre_MOP
            sops_anciens,    # H: SOPs_Anciens
            sops_nouveaux_str, # I: SOPs_Nouveaux
            sops_refs,       # J: SOPs_Refs_OPSSD
            sops_titres,     # K: Titres_SOPs
            nb_sops,         # L: Nb_SOPs
        ])
    return rows


# ============================================================
# STYLES
# ============================================================

BLUE_DARK   = '1F3864'
WHITE       = 'FFFFFF'
GREEN_LIGHT = 'C6EFCE'
VIOLET_LIGHT= 'D9B8FF'
ORANGE      = 'F4B942'
GRAY_LIGHT  = 'F2F2F2'
BLUE_LIGHT  = 'DCE6F1'
YELLOW_LIGHT= 'FFFF99'

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def bold_font(size=11, color='000000', name='Calibri'):
    return Font(bold=True, size=size, color=color, name=name)

def normal_font(size=10, color='000000', name='Calibri'):
    return Font(bold=False, size=size, color=color, name=name)

def center_align(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

thin = Side(style='thin', color='AAAAAA')
thick = Side(style='medium', color='1F3864')

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)


# ============================================================
# SHEET BUILDERS
# ============================================================

def build_recherche(ws):
    ws.title = "Recherche"

    # ── Largeurs colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 55

    # ── Titre principal (A1:G1)
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = "RECHERCHE MOP / SOP — NOC Telehouse"
    c.fill = fill(BLUE_DARK)
    c.font = Font(bold=True, size=16, color=WHITE, name='Calibri')
    c.alignment = center_align()
    ws.row_dimensions[1].height = 36

    # ── Zone de recherche (fond bleu clair A2:G8)
    for row in range(2, 9):
        ws.row_dimensions[row].height = 24
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill(BLUE_LIGHT)

    # Labels
    ws['A3'].value = "Alarme / Mot-clé :"
    ws['A3'].font = bold_font(11, BLUE_DARK)
    ws['A3'].alignment = left_align()

    ws['A5'].value = "Domaine :"
    ws['A5'].font = bold_font(11, BLUE_DARK)
    ws['A5'].alignment = left_align()

    ws['A7'].value = "Criticité :"
    ws['A7'].font = bold_font(11, BLUE_DARK)
    ws['A7'].alignment = left_align()

    # Champs de saisie
    ws['B3'].value = ""
    ws['B3'].fill = fill(WHITE)
    ws['B3'].border = thin_border()
    ws['B3'].alignment = left_align()
    ws.merge_cells('B3:G3')

    ws['B5'].value = "Tous"
    ws['B5'].fill = fill(WHITE)
    ws['B5'].border = thin_border()
    ws['B5'].alignment = left_align()

    ws['B7'].value = "Tous"
    ws['B7'].fill = fill(WHITE)
    ws['B7'].border = thin_border()
    ws['B7'].alignment = left_align()

    # Validation déroulante Domaine
    domaines = '"Tous,SUPERVISION,PHOTONIQUE,CARTE SERVICE,CARTE MANAGEMENT,CARTE INFRA,CHASSIS,PORT NNI,PORT UNI"'
    dv_dom = DataValidation(type='list', formula1=domaines, allow_blank=True, showDropDown=False)
    dv_dom.sqref = 'B5'
    ws.add_data_validation(dv_dom)

    # Validation déroulante Criticité
    criticites = '"Tous,MINEUR,MAJEUR,CRITIQUE"'
    dv_crit = DataValidation(type='list', formula1=criticites, allow_blank=True, showDropDown=False)
    dv_crit.sqref = 'B7'
    ws.add_data_validation(dv_crit)

    # ── Note sous la zone de recherche
    ws.merge_cells('A8:G8')
    ws['A8'].value = "Saisissez un mot-clé (alarme, équipement...) et/ou sélectionnez un domaine et une criticité."
    ws['A8'].font = Font(italic=True, size=9, color='666666', name='Calibri')
    ws['A8'].alignment = left_align()
    ws['A8'].fill = fill(BLUE_LIGHT)

    # ── Séparateur
    ws.row_dimensions[9].height = 8

    # ── En-têtes résultats (row 10)
    headers = ['ID', 'Alarme Lightsoft / Symptôme', 'Domaine', 'Criticité', 'MOP de référence', 'SOPs à appliquer (nouveaux IDs)', 'Titres SOPs']
    ws.row_dimensions[10].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=col)
        c.value = h
        c.fill = fill(BLUE_DARK)
        c.font = bold_font(10, WHITE)
        c.alignment = center_align(wrap=True)
        c.border = thin_border()

    # ── Formule FILTER (A11) — spill vers le bas et la droite
    # DB range: A2:L41 (40 incidents)
    # Colonnes retournées: A(ID), B(symptome), C(domaine), D(criticite), F(ref_mop), I(sops_nouveaux), K(titres_sops)
    # On filtre sur range complet B:L et on sélectionne colonnes via CHOOSECOLS si dispo
    # Fallback: on filtre A:L et affiche en entier (7 colonnes utiles = A,B,C,D,F,I,K → index 1,2,3,4,6,9,11)

    nb_rows = len(INCIDENTS)
    last_row = 1 + nb_rows  # row 1=header in DB, so data rows 2 to nb_rows+1

    # FILTER avec condition texte (insensible casse) + domaine + criticite
    # On utilise CHOOSECOLS pour sélectionner exactement les colonnes voulues
    filter_range = f"DB!A2:L{last_row}"
    col_a = f"DB!A2:A{last_row}"
    col_b = f"DB!B2:B{last_row}"
    col_c = f"DB!C2:C{last_row}"
    col_d = f"DB!D2:D{last_row}"

    cond_text  = f'IF($B$3="",1,ISNUMBER(SEARCH($B$3,{col_b})))'
    cond_dom   = f'IF($B$5="Tous",1,{col_c}=$B$5)'
    cond_crit  = f'IF($B$7="Tous",1,{col_d}=$B$7)'
    conditions = f'({col_a}<>"")*{cond_text}*{cond_dom}*{cond_crit}'

    # CHOOSECOLS(filtered, 1,2,3,4,6,9,11) = ID,Symptome,Domaine,Criticite,Ref_MOP,SOPs_Nouveaux,Titres_SOPs
    formula = (
        f'=IFERROR('
        f'CHOOSECOLS(FILTER({filter_range},{conditions}),1,2,3,4,6,9,11),'
        f'{{"Aucun résultat — vérifier le mot-clé ou sélectionner Tous","","","","","",""}}'
        f')'
    )
    ws['A11'].value = formula
    ws['A11'].alignment = left_align(wrap=True)

    # Style conditionnel simulé : colorer les lignes résultats (A11:G60)
    # On applique un fond clair pour les zones résultats attendues
    for row in range(11, 61):
        ws.row_dimensions[row].height = 22
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.alignment = left_align(wrap=True)
            c.font = normal_font(9)
            if col == 5:  # MOP ref → vert clair
                c.fill = fill(GREEN_LIGHT)
            elif col in (6, 7):  # SOPs → violet clair
                c.fill = fill(VIOLET_LIGHT)
            else:
                c.fill = fill(GRAY_LIGHT)


def build_db(ws):
    ws.title = "DB"

    # Largeurs
    widths = [8, 60, 20, 12, 8, 30, 35, 40, 30, 30, 65, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # En-têtes
    headers = [
        'ID', 'Alarme_Lightsoft', 'Domaine', 'Criticite',
        'Famille', 'Ref_MOP', 'Titre_MOP',
        'SOPs_Anciens', 'SOPs_Nouveaux', 'SOPs_Refs_OPSSD', 'Titres_SOPs', 'Nb_SOPs'
    ]
    ws.row_dimensions[1].height = 24
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.fill = fill(BLUE_DARK)
        c.font = bold_font(10, WHITE)
        c.alignment = center_align()
        c.border = thin_border()

    # Données
    rows = build_db_rows()
    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.alignment = left_align(wrap=False)
            c.font = normal_font(9)
            c.border = thin_border()
            # Colorier criticité
            if col_idx == 4:
                color_map = {'MINEUR': 'C6EFCE', 'MAJEUR': 'FFEB9C', 'CRITIQUE': 'FFC7CE'}
                c.fill = fill(color_map.get(val, 'FFFFFF'))
            elif col_idx == 6:
                c.fill = fill(GREEN_LIGHT)
            elif col_idx in (9, 10):
                c.fill = fill(VIOLET_LIGHT)


def build_legende(ws):
    ws.title = "Légende SOPs"

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 55

    # Titre
    ws.merge_cells('A1:E1')
    ws['A1'].value = "LÉGENDE — 9 SOPs cibles (après fusion 13→9)"
    ws['A1'].fill = fill(BLUE_DARK)
    ws['A1'].font = bold_font(13, WHITE)
    ws['A1'].alignment = center_align()
    ws.row_dimensions[1].height = 30

    # En-têtes
    headers = ['Nouvel ID', 'Réf OPSSD', 'Titre', 'Domaine', 'Absorbe (anciens IDs)']
    ws.row_dimensions[2].height = 24
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.fill = fill(BLUE_DARK)
        c.font = bold_font(10, WHITE)
        c.alignment = center_align()
        c.border = thin_border()

    # Construire le mapping inverse (quels anciens IDs → ce nouvel ID)
    absorbe_map = {}
    for old, new in MAPPING_OLD_NEW.items():
        if old != new:
            absorbe_map.setdefault(new, []).append(old)

    domaines_sop = {
        'SOP-01': 'Accès / OOB',
        'SOP-02': 'CLI Ribbon',
        'SOP-03': 'Escalade TAC',
        'SOP-04': 'Signal OTN',
        'SOP-05': 'Performance',
        'SOP-06': 'Hardware',
        'SOP-07': 'Protection OLP',
        'SOP-08': 'Validation',
        'SOP-09': 'Firewall mgmt',
    }

    sop_colors = [VIOLET_LIGHT, 'E8F5E9', 'FFF9C4', 'E3F2FD', 'FCE4EC',
                  'FFF3E0', 'E8EAF6', 'E0F7FA', 'F3E5F5']

    for row_idx, (sop_id, (ref, titre)) in enumerate(SOP_NEW.items(), 3):
        ws.row_dimensions[row_idx].height = 22
        absorbe = ', '.join(sorted(absorbe_map.get(sop_id, [])))
        row_data = [sop_id, ref, titre, domaines_sop[sop_id], absorbe or '—']
        row_color = sop_colors[(row_idx - 3) % len(sop_colors)]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.fill = fill(row_color)
            c.font = bold_font(10) if col_idx == 1 else normal_font(10)
            c.alignment = left_align(wrap=True)
            c.border = thin_border()

    # Mapping table ancien→nouveau
    ws.row_dimensions[14].height = 28
    ws.merge_cells('A14:E14')
    ws['A14'].value = "MAPPING COMPLET — Anciens IDs → Nouveaux IDs"
    ws['A14'].fill = fill(BLUE_DARK)
    ws['A14'].font = bold_font(11, WHITE)
    ws['A14'].alignment = center_align()

    ws.row_dimensions[15].height = 20
    for col, h in enumerate(['Ancien ID', 'Nouveau ID', 'Titre nouveau', '', ''], 1):
        c = ws.cell(row=15, column=col)
        c.value = h
        c.fill = fill(GRAY_LIGHT)
        c.font = bold_font(9)
        c.alignment = center_align()
        c.border = thin_border()

    for row_idx, (old, new) in enumerate(sorted(MAPPING_OLD_NEW.items()), 16):
        ws.row_dimensions[row_idx].height = 18
        titre_new = SOP_NEW[new][1]
        note = 'fusionné' if old != new else '—'
        for col_idx, val in enumerate([old, new, titre_new, note, ''], 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.font = normal_font(9)
            c.border = thin_border()
            if col_idx == 1:
                c.fill = fill('FFF9C4') if old != new else fill('FFFFFF')
            elif col_idx == 2:
                c.fill = fill(VIOLET_LIGHT)


def build_mode_emploi(ws):
    ws.title = "Mode d'emploi"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80

    ws.merge_cells('A1:B1')
    ws['A1'].value = "MODE D'EMPLOI — Moteur de recherche MOP/SOP NOC"
    ws['A1'].fill = fill(BLUE_DARK)
    ws['A1'].font = bold_font(14, WHITE)
    ws['A1'].alignment = center_align()
    ws.row_dimensions[1].height = 36

    instructions = [
        ("1.", "Onglet 'Recherche' — Saisissez un mot-clé dans la cellule B3 (alarme Lightsoft, équipement, symptôme...)."),
        ("2.", "Sélectionnez un Domaine dans B5 (ou laissez 'Tous' pour chercher dans tous les domaines)."),
        ("3.", "Sélectionnez une Criticité dans B7 (ou laissez 'Tous')."),
        ("4.", "Les résultats s'affichent automatiquement — MOP de référence (vert) + SOPs à appliquer (violet)."),
        ("5.", "Si 'Aucun résultat' : vérifier l'orthographe ou élargir les filtres en sélectionnant 'Tous'."),
        ("", ""),
        ("NOTE", "Ce fichier requiert Excel 365 ou Excel 2021 pour les formules FILTER/CHOOSECOLS."),
        ("NOTE", "LibreOffice Calc ne supporte pas FILTER() — utiliser Excel Online (SharePoint) ou Excel desktop."),
        ("", ""),
        ("LÉGENDE", "Onglet 'Légende SOPs' : détail des 9 SOPs cibles + mapping anciens→nouveaux IDs."),
        ("DONNÉES", "Onglet 'DB' (masqué en production) : 40 incidents source de vérité."),
        ("", ""),
        ("Généré", "2026-04-14 — Source : mop-incidents-alignment.yml + mop-route.json + mop-get.json"),
        ("Version", "V1.0 — Plan de fusion 13→9 SOPs (sop-rationalisation-2026-04-13.md)"),
    ]

    for row_idx, (num, text) in enumerate(instructions, 3):
        ws.row_dimensions[row_idx].height = 22
        ws.cell(row=row_idx, column=1).value = num
        ws.cell(row=row_idx, column=1).font = bold_font(10, BLUE_DARK)
        ws.cell(row=row_idx, column=1).alignment = center_align()
        ws.cell(row=row_idx, column=2).value = text
        ws.cell(row=row_idx, column=2).font = normal_font(10)
        ws.cell(row=row_idx, column=2).alignment = left_align(wrap=True)
        if num.startswith("NOTE"):
            ws.cell(row=row_idx, column=1).fill = fill(ORANGE)
            ws.cell(row=row_idx, column=2).fill = fill('FFF3CD')
        elif num.isdigit():
            ws.cell(row=row_idx, column=2).fill = fill(GRAY_LIGHT)


# ============================================================
# MAIN
# ============================================================

def main():
    output_path = Path('/home/mobuone/VPAI/docs/mop-sop-search.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Supprimer la feuille par défaut
    del wb['Sheet']

    # Créer les 4 onglets
    ws_rech = wb.create_sheet("Recherche")
    ws_db   = wb.create_sheet("DB")
    ws_leg  = wb.create_sheet("Légende SOPs")
    ws_mode = wb.create_sheet("Mode d'emploi")

    build_recherche(ws_rech)
    build_db(ws_db)
    build_legende(ws_leg)
    build_mode_emploi(ws_mode)

    # Onglet actif = Recherche (index 0)
    wb.active = ws_rech

    # Masquer l'onglet DB en production (tab hidden)
    # ws_db.sheet_state = 'hidden'  # décommenter pour production

    wb.save(output_path)
    print(f"[OK] Fichier généré : {output_path}")

    # Vérification rapide
    wb2 = openpyxl.load_workbook(output_path)
    db = wb2['DB']
    n_rows = sum(1 for row in db.iter_rows(min_row=2, values_only=True) if row[0])
    print(f"[OK] DB : {n_rows} incidents chargés")
    rech = wb2['Recherche']
    formula_cell = rech['A11'].value
    if formula_cell and formula_cell.startswith('='):
        print(f"[OK] Formule FILTER présente en A11")
    else:
        print(f"[WARN] Formule FILTER absente ou vide en A11: {formula_cell!r}")
    print(f"[OK] Onglets : {wb2.sheetnames}")


if __name__ == '__main__':
    main()
