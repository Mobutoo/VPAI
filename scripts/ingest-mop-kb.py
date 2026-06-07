#!/usr/bin/env python3
"""Ingest documents from DUFS /Documentation into mop_kb (768d via mop-embed).

Pipeline:
  PDF  → base64 → Gemini 2.5 Pro multimodal (mop-ingest) → markdown
  DOCX → python-docx text extraction → Gemini text-only → markdown
  XLSX → openpyxl text extraction → Gemini text-only → markdown
  markdown → chunk 1600/200 → mop-embed (768d) → mop_kb upsert

Usage:
  pip install pymupdf python-docx openpyxl   # first time only
  python3 scripts/ingest-mop-kb.py
"""

import base64, hashlib, json, os, re, sys, time, urllib.request, urllib.error, urllib.parse

LITELLM_URL  = os.environ.get("LITELLM_URL", "https://llm.ewutelo.cloud")
LITELLM_KEY  = os.environ.get("LITELLM_KEY", "sk-lm-08041985e12102015m")
QDRANT_URL   = os.environ.get("QDRANT_URL",  "https://qd.ewutelo.cloud")
QDRANT_KEY   = os.environ.get("QDRANT_KEY",  "sk-qd-08041985e12102015m")
DUFS_URL     = os.environ.get("DUFS_URL",    "https://drop.ewutelo.cloud")
DUFS_FOLDER  = os.environ.get("DUFS_FOLDER", "/Documentation")
COLLECTION   = "mop_kb"
EMBED_MODEL  = "mop-embed"
INGEST_MODEL = "mop-ingest"
CHUNK_MAX    = 1600
CHUNK_OVL    = 200
EMBED_BATCH  = 20

PROMPTS = {
    "pdf": (
        "Tu es un expert en systèmes télécom (DWDM, SDH, IP/MPLS). "
        "Extrais la totalité de ce document technique NOC en markdown structuré. "
        "Respecte la hiérarchie: H1 titre, H2 sections, H3 sous-sections. "
        "TABLEAUX: reproduis chaque tableau en markdown complet avec toutes les colonnes et lignes — "
        "ne tronque jamais un tableau, inclus l'en-tête et la ligne de séparation `|---|`. "
        "IMAGES ET SCHÉMAS: pour chaque image, diagramme, schéma, figure ou capture d'écran, "
        "génère une description textuelle détaillée sous forme de paragraphe préfixé par '[IMAGE: ...]' "
        "incluant: type de schéma, composants visibles, connexions, labels, valeurs numériques, "
        "flux de données ou alarmes représentés. "
        "Procédures en listes numérotées. "
        "Conserve TOUS les codes alarme, références SP, compteurs de performance, seuils et commandes CLI. "
        "Retourne UNIQUEMENT le markdown, sans préambule ni commentaire."
    ),
    "docx": (
        "Tu es un expert en systèmes télécom (DWDM, SDH, IP/MPLS). "
        "Voici le texte extrait d'un document Word. Restructure-le en markdown: "
        "H1 titre principal, H2 chapitres, H3 sous-sections. "
        "TABLEAUX: reproduis chaque tableau en markdown complet avec toutes les colonnes et lignes — "
        "ne tronque jamais un tableau, inclus l'en-tête et la ligne de séparation `|---|`. "
        "Procédures en listes numérotées. "
        "Conserve TOUS les codes alarme, références SP, compteurs de performance, seuils et commandes CLI. "
        "Retourne UNIQUEMENT le markdown, sans préambule ni commentaire."
    ),
    "xlsx": (
        "Tu es un expert en systèmes télécom (DWDM, SDH, IP/MPLS). "
        "Voici le contenu tabulaire extrait d'un fichier Excel. "
        "Chaque ligne est un incident ou une procédure NOC. "
        "Convertis en markdown: H2 par feuille, tableau markdown complet par section. "
        "TABLEAUX: inclus toujours l'en-tête avec ligne de séparation `|---|` et toutes les lignes. "
        "Conserve TOUS les codes alarme, références SP, compteurs de performance, seuils et commandes CLI. "
        "Retourne UNIQUEMENT le markdown, sans préambule ni commentaire."
    ),
}

# ── HTTP ──────────────────────────────────────────────────────────────────────

def http_get(url: str) -> bytes:
    with urllib.request.urlopen(urllib.request.Request(url), timeout=120) as r:
        return r.read()

def http_post(url: str, body: bytes, headers: dict) -> dict:
    req = urllib.request.Request(url, data=body, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=300) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code}: {e.read().decode('utf-8','replace')[:300]}")

# ── DUFS ──────────────────────────────────────────────────────────────────────

def list_dufs() -> list[dict]:
    data = json.loads(http_get(f"{DUFS_URL}{DUFS_FOLDER}/?json"))
    return [p for p in data.get("paths", []) if p["path_type"] == "File"]

def download(filename: str) -> bytes:
    url = f"{DUFS_URL}{DUFS_FOLDER}/{urllib.parse.quote(filename)}"
    print(f"  ↓ {filename} ...", end=" ", flush=True)
    data = http_get(url)
    print(f"{len(data)//1024} KB")
    return data

# ── Text extraction (DOCX / XLSX) ─────────────────────────────────────────────

def extract_docx(data: bytes) -> str:
    import io
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        raise RuntimeError("python-docx manquant: pip install python-docx")

def extract_xlsx(data: bytes) -> str:
    import io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"## {sheet}\n\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except ImportError:
        raise RuntimeError("openpyxl manquant: pip install openpyxl")

# ── Gemini via LiteLLM ────────────────────────────────────────────────────────

def to_markdown(file_bytes: bytes, filename: str, ext: str) -> str:
    prompt = PROMPTS.get(ext, PROMPTS["pdf"])

    if ext == "pdf":
        b64 = base64.b64encode(file_bytes).decode()
        content = [
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{b64}"}},
        ]
    else:
        if ext == "docx":
            text = extract_docx(file_bytes)
        else:
            text = extract_xlsx(file_bytes)
        content = [{"type": "text", "text": prompt + "\n\n---\n\n" + text[:60000]}]

    print(f"  → Gemini ({ext}, {len(file_bytes)//1024}KB)...", end=" ", flush=True)
    result = http_post(
        f"{LITELLM_URL}/v1/chat/completions",
        json.dumps({"model": INGEST_MODEL, "messages": [{"role": "user", "content": content}], "max_tokens": 8192}).encode(),
        {"Content-Type": "application/json", "Authorization": f"Bearer {LITELLM_KEY}"},
    )
    md = result["choices"][0]["message"]["content"]
    print(f"{len(md)} chars")
    return md

# ── Chunking ──────────────────────────────────────────────────────────────────

def _split_paragraphs(text: str) -> list[str]:
    """Split text into paragraphs, keeping markdown tables as atomic blocks."""
    paragraphs = []
    table_lines: list[str] = []
    in_table = False
    buf_lines: list[str] = []

    for line in text.split("\n"):
        is_table_row = line.strip().startswith("|")
        if is_table_row:
            if not in_table:
                # flush pending paragraph
                if buf_lines:
                    block = "\n".join(buf_lines).strip()
                    if block:
                        paragraphs.append(block)
                    buf_lines = []
                in_table = True
            table_lines.append(line)
        else:
            if in_table:
                # flush completed table as one atomic paragraph
                paragraphs.append("\n".join(table_lines))
                table_lines = []
                in_table = False
            if line.strip() == "":
                # blank line = paragraph separator
                block = "\n".join(buf_lines).strip()
                if block:
                    paragraphs.append(block)
                buf_lines = []
            else:
                buf_lines.append(line)

    if in_table and table_lines:
        paragraphs.append("\n".join(table_lines))
    elif buf_lines:
        block = "\n".join(buf_lines).strip()
        if block:
            paragraphs.append(block)

    return [p for p in paragraphs if p.strip()]


def chunk(text: str) -> list[str]:
    chunks = []
    for sec in re.split(r'(?=\n#{1,3} )', text):
        sec = sec.strip()
        if not sec:
            continue
        if len(sec) <= CHUNK_MAX:
            chunks.append(sec)
            continue
        buf = ""
        for p in _split_paragraphs(sec):
            if not p:
                continue
            cand = (buf + "\n\n" + p).strip() if buf else p
            if len(cand) <= CHUNK_MAX:
                buf = cand
            else:
                if buf:
                    chunks.append(buf)
                if len(p) <= CHUNK_MAX:
                    buf = p
                else:
                    # Oversized atomic block (large table or paragraph): keep whole
                    chunks.append(p)
                    buf = ""
        if buf:
            chunks.append(buf)
    return [c for c in chunks if len(c) >= 20]

# ── Embedding ─────────────────────────────────────────────────────────────────

def embed_batch(texts: list[str]) -> list[list[float]]:
    result = http_post(
        f"{LITELLM_URL}/v1/embeddings",
        json.dumps({"model": EMBED_MODEL, "input": texts}).encode(),
        {"Content-Type": "application/json", "Authorization": f"Bearer {LITELLM_KEY}"},
    )
    return [x["embedding"] for x in sorted(result["data"], key=lambda x: x["index"])]

# ── Qdrant ────────────────────────────────────────────────────────────────────

def stable_uuid(s: str) -> str:
    h = hashlib.sha256(s.encode()).hexdigest()
    return f"{h[:8]}-{h[8:12]}-4{h[13:16]}-{h[16:20]}-{h[20:32]}"

def delete_by_filename(filename: str) -> None:
    body = json.dumps({"filter": {"must": [{"key": "filename", "match": {"value": filename}}]}}).encode()
    req = urllib.request.Request(
        f"{QDRANT_URL}/collections/{COLLECTION}/points/delete?wait=true",
        data=body, method="POST",
        headers={"Content-Type": "application/json", "api-key": QDRANT_KEY},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        res = json.loads(r.read())
    print(f"  ✗ Cleared old points: {res.get('result',{}).get('status','?')}")

def upsert_points(points: list[dict]) -> None:
    body = json.dumps({"points": points}).encode()
    req = urllib.request.Request(
        f"{QDRANT_URL}/collections/{COLLECTION}/points?wait=true",
        data=body, method="PUT",
        headers={"Content-Type": "application/json", "api-key": QDRANT_KEY},
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        res = json.loads(r.read())
    if res.get("status") != "ok":
        raise RuntimeError(f"Qdrant upsert error: {res}")

# ── Main ──────────────────────────────────────────────────────────────────────

def process(filename: str, file_bytes: bytes) -> int:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ("pdf", "docx", "xlsx", "xls"):
        print(f"  ⚠ Format non supporté: {ext} — skip")
        return 0

    markdown = to_markdown(file_bytes, filename, ext if ext != "xls" else "xlsx")
    chunks = chunk(markdown)
    print(f"  → {len(chunks)} chunks")
    if not chunks:
        return 0

    delete_by_filename(filename)

    all_vectors = []
    for i in range(0, len(chunks), EMBED_BATCH):
        batch = chunks[i:i + EMBED_BATCH]
        print(f"  → Embed {i+1}-{min(i+EMBED_BATCH, len(chunks))}/{len(chunks)}...", end=" ", flush=True)
        vecs = embed_batch(batch)
        all_vectors.extend(vecs)
        print("ok")
        time.sleep(0.1)

    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    points = []
    for i, (c, v) in enumerate(zip(chunks, all_vectors)):
        m = re.search(r'^#{1,3} (.+)', c, re.MULTILINE)
        points.append({
            "id": stable_uuid(f"{filename}:{i}:{c[:80]}"),
            "vector": v,
            "payload": {
                "filename":    filename,
                "source_kind": "reference",
                "category":    "guide",
                "namespace":   "mop",
                "language":    "fr",
                "doc_type":    ext,
                "section":     m.group(1).strip() if m else "",
                "chunk_index": i,
                "chunk_total": len(chunks),
                "text":        c,
                "indexed_at":  now,
            },
        })

    for i in range(0, len(points), 100):
        upsert_points(points[i:i + 100])
    print(f"  ✓ {len(points)} points → {COLLECTION}")
    return len(points)

def main():
    print(f"=== MOP KB Ingestor ===")
    print(f"DUFS:    {DUFS_URL}{DUFS_FOLDER}")
    print(f"Qdrant:  {QDRANT_URL} / {COLLECTION}")
    print()

    files = list_dufs()
    print(f"{len(files)} fichiers trouvés:")
    for f in files:
        print(f"  - {f['name']} ({f['size']//1024} KB)")
    print()

    total = 0
    for i, f in enumerate(files):
        print(f"\n[{i+1}/{len(files)}] {f['name']}")
        try:
            data = download(f["name"])
            n = process(f["name"], data)
            total += n
        except Exception as e:
            print(f"  ✗ ERREUR: {e}", file=sys.stderr)
            import traceback; traceback.print_exc()

    print(f"\n=== Terminé: {total} points dans {COLLECTION} ===")

if __name__ == "__main__":
    main()
